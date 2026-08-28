import streamlit as st
import os
import re
import pandas as pd
from io import BytesIO
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# ==========================================
# 1. CẤU HÌNH TRANG CHỦ ĐẠO
# ==========================================
st.set_page_config(page_title="Phần mềm Quản lý Suất ăn Nhà máy", page_icon="🍱", layout="centered")

# ==========================================
# 2. HÀM KẾT NỐI VÀ ĐẨY DỮ LIỆU LÊN GOOGLE SHEETS
# ==========================================
def push_to_google_sheets(df_summary, df_vang, df_tang, total_final):
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        
        # Lấy thông tin credentials từ st.secrets
        creds_dict = dict(st.secrets["GOOGLE_CREDENTIALS"])
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        
        sheet_id = st.secrets["SHEETS_URL"]
        spreadsheet = client.open_by_key(sheet_id)
        
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        date_str = datetime.now().strftime("%d/%m/%Y")

        # 1. Đẩy Tổng hợp suất ăn
        try:
            ws_summary = spreadsheet.worksheet("Lich_Su_Tong_Hop")
        except gspread.exceptions.WorksheetNotFound:
            ws_summary = spreadsheet.add_worksheet(title="Lich_Su_Tong_Hop", rows="1000", cols="10")
            ws_summary.append_row(["Thời gian chốt", "Ngày", "Bộ Phận", "Tổng NS", "Cắt Tháng", "Vắng Hôm Nay", "Đột Xuất Giảm", "Khách Tăng", "THỰC TẾ ĐẶT"])

        for _, row in df_summary.iterrows():
            ws_summary.append_row([
                now_str, date_str, row["Bộ Phận"], int(row["Tổng NS"]), 
                int(row["Cắt Tháng (-)"]), int(row["Vắng Hôm Nay (-)"]), 
                int(row["Đột Xuất Giảm (-)"]), int(row["Khách Tăng (+)"]), int(row["THỰC TẾ ĐẶT"])
            ])

        # 2. Đẩy Chi tiết Vắng mặt
        if not df_vang.empty:
            try:
                ws_vang = spreadsheet.worksheet("Lich_Su_Vang_Mat")
            except gspread.exceptions.WorksheetNotFound:
                ws_vang = spreadsheet.add_worksheet(title="Lich_Su_Vang_Mat", rows="1000", cols="10")
                ws_vang.append_row(["Thời gian", "Ngày", "Mã NV (Code)", "Họ và Tên", "Bộ Phận", "Lý do vắng"])

            for _, row in df_vang.iterrows():
                ws_vang.append_row([
                    now_str, date_str, str(row.get("mã nv (code)", row.get("code", ""))), 
                    str(row.get("họ và tên", "")), str(row.get("bộ phận", row.get("bp", ""))), 
                    str(row.get("lý do vắng", "Nghỉ/Cắt đột xuất"))
                ])

        # 3. Đẩy Chi tiết Khách
        if not df_tang.empty:
            try:
                ws_tang = spreadsheet.worksheet("Lich_Su_Khach")
            except gspread.exceptions.WorksheetNotFound:
                ws_tang = spreadsheet.add_worksheet(title="Lich_Su_Khach", rows="1000", cols="10")
                ws_tang.append_row(["Thời gian", "Ngày", "Đoàn khách/Người", "Bộ Phận đón", "Số lượng", "Ghi chú"])

            for _, row in df_tang.iterrows():
                ws_tang.append_row([
                    now_str, date_str, str(row["Đoàn khách/Người"]), 
                    str(row["Bộ Phận đón"]), int(row["Số lượng"]), str(row["Ghi chú"])
                ])
                
        return True
    except Exception as e:
        st.error(f"Lỗi đồng bộ Google Sheets: {e}")
        return False

# ==========================================
# 3. GIAO DIỆN CHÍNH CỦA APP
# ==========================================
def create_empty_templates():
    wb = openpyxl.Workbook()
    
    ws_tong = wb.active
    ws_tong.title = "DS_Tong_Nhan_Su"
    ws_tong.append(["Mã NV (Code)", "Họ và Tên", "Bộ Phận"])
    ws_tong.append(["0319", "Đặng Chí Hiếu", "PNV"])
    ws_tong.append(["0546", "Trần Mỹ Vân", "BPB"])
    
    ws_cat = wb.create_sheet(title="DS_Cat_Com_Thang")
    ws_cat.append(["STT", "BP", "Code", "Họ và Tên", "Ghi chú"])
    ws_cat.append(["1", "PNV", "0319", "Đặng Chí Hiếu", "Mang cơm nhà"])
    
    ws_vang = wb.create_sheet(title="HDSD_Huong_Dan")
    ws_vang.append(["HƯỚNG DẪN: Mỗi ngày tạo 1 Sheet mới đặt tên theo định dạng DD-MM (Ví dụ: 28-08)"])
    
    ws_ngay = wb.create_sheet(title="28-08")
    ws_ngay.append(["Mã NV (Code)", "Họ và Tên", "Bộ Phận", "Lý do vắng"])
    ws_ngay.append(["0546", "Trần Mỹ Vân", "BPB", "Nghỉ phép"])
    
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

st.title("🍱 Quản lý & Tổng hợp Suất ăn Nhà máy")
st.markdown("Hệ thống tự động trừ hao cắt cơm tháng, bắt lỗi trùng lặp và đồng bộ dữ liệu lên đám mây phục vụ tra cứu.")

# 1. Tải file mẫu
with st.expander("📥 Tải File Excel Mẫu (Template) cho 4 Phòng ban"):
    st.markdown("File mẫu có sẵn hướng dẫn và quy định đặt tên Sheet theo ngày (`DD-MM`).")
    st.download_button(
        label="⬇️ Tải File Mẫu Chuẩn (.xlsx)",
        data=create_empty_templates(),
        file_name="Template_Quan_Ly_Suat_An.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.markdown("---")

# 2. Upload file đầu vào
st.subheader("1. Nạp Dữ Liệu Đầu Vào")
col1, col2 = st.columns(2)
with col1:
    file_tong = st.file_uploader("Tải lên File 'Danh sách Tổng & Cắt Cơm Tháng'", type=["xlsx", "xls"])
with col2:
    file_vangs = st.file_uploader("Tải lên 4 File 'Báo vắng hôm nay' của các phòng ban (Kéo thả cùng lúc)", type=["xlsx", "xls"], accept_multiple_files=True)

st.subheader("2. Điều chỉnh Đột Xuất (Trực tiếp trên App)")
st.markdown("🔻 **Người nghỉ ngang / Cắt cơm đột xuất phát sinh ngoài danh sách:**")
df_giam_init = pd.DataFrame([{"Mã NV (Code)": "", "Họ và Tên": "", "Bộ Phận": "", "Lý do": ""}] * 2)
edited_df_giam = st.data_editor(df_giam_init, num_rows="dynamic", key="editor_giam", use_container_width=True)

st.markdown("🔺 **Khách đối tác / Tăng cường (Tăng suất):**")
df_tang_init = pd.DataFrame([{"Đoàn khách/Người": "", "Bộ Phận đón": "", "Số lượng": 0, "Ghi chú": ""}] * 2)
edited_df_tang = st.data_editor(df_tang_init, num_rows="dynamic", key="editor_tang", use_container_width=True)

if st.button("📊 Chốt Số Lượng & Đẩy Lên Google Sheets", type="primary"):
    if not file_tong:
        st.error("⚠️ Vui lòng tải File Danh sách Tổng để hệ thống bắt đầu tính toán.")
    else:
        with st.spinner("🤖 Đang quét dữ liệu, lọc chống trùng lặp và đồng bộ đám mây..."):
            try:
                # Đọc File Tổng & Cắt tháng
                xls_tong = pd.ExcelFile(file_tong)
                df_tong = pd.read_excel(xls_tong, sheet_name=0) 
                df_cat_thang = pd.read_excel(xls_tong, sheet_name=1) 
                
                def normalize_cols(df):
                    df.columns = [str(c).strip().lower() for c in df.columns]
                    bp_col = next((c for c in df.columns if 'bộ phận' in c or c == 'bp'), None)
                    code_col = next((c for c in df.columns if 'code' in c or 'mã' in c), None)
                    return df, bp_col, code_col

                df_tong, bp_col_tong, code_col_tong = normalize_cols(df_tong)
                df_cat_thang, bp_col_cat, code_col_cat = normalize_cols(df_cat_thang)

                # Lấy danh sách mã nhân viên cắt cơm tháng (Blacklist) để chống trùng lặp
                blacklist_codes = set()
                if code_col_cat and not df_cat_thang.empty:
                    blacklist_codes = set(df_cat_thang[code_col_cat].astype(str).str.strip().str.upper())

                # Xử lý 4 file báo vắng của phòng ban (Tự động dò sheet theo ngày hiện tại)
                current_day_str = datetime.now().strftime("%d-%m") # Lấy dạng "28-08"
                current_day_alt = datetime.now().strftime("%d/%m")
                
                all_vang_rows = []
                if file_vangs:
                    for fv in file_vangs:
                        try:
                            xls_sub = pd.ExcelFile(fv)
                            sheet_names = xls_sub.sheet_names
                            
                            # Tìm sheet khớp ngày hiện tại hoặc lấy sheet cuối cùng
                            target_sheet = sheet_names[-1]
                            for s in sheet_names:
                                if current_day_str in str(s) or current_day_alt in str(s):
                                    target_sheet = s
                                    break
                            
                            df_sub = pd.read_excel(xls_sub, sheet_name=target_sheet)
                            df_sub_norm, bp_s, code_s = normalize_cols(df_sub)
                            
                            if code_s:
                                for _, r in df_sub_norm.iterrows():
                                    emp_code = str(r[code_s]).strip().upper()
                                    # CHỐNG TRÙNG LẶP: Nếu đã cắt cơm tháng rồi thì bỏ qua không trừ đúp
                                    if emp_code not in blacklist_codes:
                                        all_vang_rows.append(r)
                        except Exception as sub_e:
                            st.warning(f"Không thể đọc file {fv.name}: {sub_e}")

                df_vang = pd.DataFrame(all_vang_rows) if all_vang_rows else pd.DataFrame()
                
                # Tính toán số liệu theo Bộ Phận
                tong_dict = df_tong.groupby(bp_col_tong).size().to_dict() if bp_col_tong else {}
                cat_dict = df_cat_thang.groupby(bp_col_cat).size().to_dict() if bp_col_cat else {}
                
                vang_dict = {}
                if not df_vang.empty:
                    _, bp_col_v, _ = normalize_cols(df_vang)
                    if bp_col_v:
                        vang_dict = df_vang.groupby(bp_col_v).size().to_dict()

                df_giam_clean = edited_df_giam[edited_df_giam["Bộ Phận"].str.strip() != ""]
                giam_dict = df_giam_clean.groupby("Bộ Phận").size().to_dict()

                df_tang_clean = edited_df_tang[(edited_df_tang["Bộ Phận đón"].str.strip() != "") & (edited_df_tang["Số lượng"] > 0)]
                tang_dict = df_tang_clean.groupby("Bộ Phận đón")["Số lượng"].sum().to_dict()

                all_bps = set(list(tong_dict.keys()) + list(cat_dict.keys()) + list(vang_dict.keys()) + list(giam_dict.keys()) + list(tang_dict.keys()))

                summary_data = []
                total_final = 0
                
                for bp in sorted(all_bps):
                    t = tong_dict.get(bp, 0)
                    c = cat_dict.get(bp, 0)
                    v = vang_dict.get(bp, 0)
                    g = giam_dict.get(bp, 0)
                    kh = tang_dict.get(bp, 0)
                    
                    thuc_te = t - c - v - g + kh
                    total_final += thuc_te
                    
                    summary_data.append({
                        "Bộ Phận": str(bp).upper(),
                        "Tổng NS": t,
                        "Cắt Tháng (-)": c,
                        "Vắng Hôm Nay (-)": v,
                        "Đột Xuất Giảm (-)": g,
                        "Khách Tăng (+)": kh,
                        "THỰC TẾ ĐẶT": thuc_te
                    })

                df_summary = pd.DataFrame(summary_data)

                # --- ĐẨY LÊN GOOGLE SHEETS ---
                push_success = push_to_google_sheets(df_summary, df_vang, df_tang_clean, total_final)
                if push_success:
                    st.success("☁️ Đã đồng bộ lịch sử thành công lên Google Sheets cho Sếp tra cứu!")

                # --- HIỂN THỊ KẾT QUẢ ---
                st.success("✅ Đã xử lý và chốt số liệu thành công!")
                st.markdown("### 📋 BẢNG CHỐT SỐ LƯỢNG HÔM NAY")
                st.dataframe(df_summary, use_container_width=True)

                # Tin nhắn Zalo
                zalo_msg = "🍱 *CHỐT SUẤT ĂN HÔM NAY:*\n\n"
                for _, row in df_summary.iterrows():
                    if row['THỰC TẾ ĐẶT'] > 0:
                        zalo_msg += f"- {row['Bộ Phận']}: {row['THỰC TẾ ĐẶT']} suất\n"
                zalo_msg += f"\n👉 **TỔNG CỘNG: {total_final} suất.**"

                st.info("💬 Copy tin nhắn này gửi Zalo nhà bếp:")
                st.code(zalo_msg, language="markdown")

            except Exception as e:
                st.error(f"Đã xảy ra lỗi hệ thống: {e}")
